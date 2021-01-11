import openpyxl
#from xlsconfig import *
import os
#import pyexcel as p
#import pandas as pd
#import numpy as np
import glob
#import shutil
import datetime
import time
import requests
from requests.auth import HTTPBasicAuth
import json


#today = date.today()
dt = datetime.datetime.today()
newtime =  time.strftime('%Y %m %d').replace(' 0', ' ')

def sendSMS(number, message):

    url = "https://secure.smslink.ro/sms/gateway/communicate/index.php"
    params = {'connection_id':'CONNECTION_ID',
           'password' : 'PASSWORD',
            'to'      : number,
            'message' : message
          }
    r = requests.get(url = url, params = params)
    print(r.text)
    print(r.status_code)



def ExcelActivities(inputXLS):
     wb_obj = openpyxl.load_workbook(inputXLS) 
     sheet_obj = wb_obj.active 
     m_row = sheet_obj.max_row 

     for i in range(2, m_row + 1): 
       nume = sheet_obj.cell(row = i, column = 1) 
       prenume = sheet_obj.cell(row = i, column = 2)
       telefon = sheet_obj.cell(row = i, column = 3)
       mesaj = sheet_obj.cell(row = i, column = 4)
       data = sheet_obj.cell(row = i, column = 5)
       #print(column1.value)
       #print(data.value)
       #print(dt.day)
       crt_day = dt.day
       crt_month = dt.month
       crt_year = dt.year
       pgm_date = data.value.split('/')
       #print(pgm_date[2])
       pgm_day = pgm_date[0]
       pgm_month = pgm_date[1]
       pgm_year = pgm_date[2]
       #print(str(crt_day) + "-" + str(pgm_day))
       #print(newtime.month)

       if int(crt_day) == int(pgm_day) and int(crt_month) == int(crt_month) and int(crt_year) == int(crt_year):
          #print("Dl/Dna " + nume.value + " este programata azi " + str(crt_day) + "/" + str(crt_month) + "/" + str(crt_year))
          customerMessage = "Stimate Dl/Dna " + nume.value + " " + prenume.value + " " + telefon.value + " " + " acesta este un test" 
          print(customerMessage) 
          sendSMS(telefon.value, customerMessage)
          time.sleep(15) 
if __name__ == "__main__":
	ExcelActivities("smsdb.xlsx")

